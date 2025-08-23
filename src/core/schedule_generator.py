"""
Schedule Generator Module - Generate Excel TA Schedules with Office Hours tracking
Extracted from generate_ta_schedule.py for integration into the main scheduler
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import re
import os
from .data_parser import get_ta_unavailable_slots


def parse_lab_assignments(labs_assigned_str):
    """
    Parse the lab assignments string to extract time slots.
    Format: "LAB D04 (Tuesday 11-14), LAB D11 (Wednesday 11-14)"
    Returns: list of tuples (day, start_hour, end_hour)
    """
    if pd.isna(labs_assigned_str) or labs_assigned_str == "None":
        return []
    
    assignments = []
    # Split by comma and process each assignment
    lab_entries = labs_assigned_str.split(", ")
    
    for entry in lab_entries:
        # Extract day and time using regex
        # Pattern: (Day start_time-end_time)
        match = re.search(r'\((\w+)\s+(\d+)-(\d+)\)', entry)
        if match:
            day = match.group(1)
            start_hour = int(match.group(2))
            end_hour = int(match.group(3))
            assignments.append((day, start_hour, end_hour))
    
    return assignments


def create_vba_instructions(ta_sheet_names):
    """
    Create VBA code instructions and files for dynamic Office Hours updates
    """
    # VBA code for Module1 (regular module)
    module_code = '''
Sub UpdateOfficeHours()
    Dim ws As Worksheet
    Dim ohWs As Worksheet
    Dim taNames As Variant
    Dim sheetNames As Variant
    Dim i As Integer, j As Integer, k As Integer
    Dim timeSlots As Variant
    Dim days As Variant
    Dim cellValue As String
    Dim ohList As String
    Dim targetCell As Range
    
    Application.ScreenUpdating = False
    
    ' Define arrays
    timeSlots = Array("8-9", "9-10", "10-11", "11-12", "12-13", "13-14", "14-15", "15-16", "16-17", "17-18", "18-19", "19-20")
    days = Array("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
    
    ' TA Names and corresponding sheet names
''' + f'''
    taNames = Array({", ".join([f'"{name}"' for name in ta_sheet_names.keys()])})
    sheetNames = Array({", ".join([f'"{sheet}"' for sheet in ta_sheet_names.values()])})
''' + '''
    
    ' Get Office Hours worksheet
    Set ohWs = ThisWorkbook.Worksheets("Office Hours")
    
    ' Clear existing Office Hours data completely
    ohWs.Range("B2:F13").ClearContents
    ohWs.Range("B2:F13").Interior.Color = xlNone
    
    ' Loop through each time slot and day
    For i = 0 To UBound(timeSlots)
        For j = 0 To UBound(days)
            ohList = ""
            
            ' Check each TA sheet for "OH" in this time slot
            For k = 0 To UBound(taNames)
                On Error Resume Next
                Set ws = ThisWorkbook.Worksheets(sheetNames(k))
                If Not ws Is Nothing Then
                    Set targetCell = ws.Cells(i + 2, j + 2) ' Offset for headers
                    cellValue = Trim(UCase(targetCell.Value))
                    
                    If cellValue = "OH" Then
                        If ohList = "" Then
                            ohList = taNames(k)
                        Else
                            ohList = ohList & ", " & taNames(k)
                        End If
                    End If
                End If
                Set ws = Nothing
                On Error GoTo 0
            Next k
            
            ' Update Office Hours sheet
            ohWs.Cells(i + 2, j + 2).Value = ohList
        Next j
    Next i
    
    Application.ScreenUpdating = True
End Sub

Sub ManualUpdateOfficeHours()
    ' This can be called manually from any cell or button
    UpdateOfficeHours
End Sub
'''

    # VBA code for ThisWorkbook (event handler)
    thisworkbook_code = '''
Private Sub Workbook_SheetChange(ByVal Sh As Object, ByVal Target As Range)
    ' Auto-update Office Hours when any cell in schedule area changes
    If Target.Count = 1 And Sh.Name <> "Office Hours" Then
        ' Check if the cell is in the schedule area (B2:F13 for most sheets)
        If Target.Row >= 2 And Target.Row <= 13 And Target.Column >= 2 And Target.Column <= 6 Then
            ' Always update Office Hours when schedule area changes
            ' This ensures both additions and deletions are captured
            Application.EnableEvents = False
            Application.DisplayAlerts = False
            Call UpdateOfficeHours
            Application.DisplayAlerts = True
            Application.EnableEvents = True
        End If
    End If
End Sub
'''
    
    return module_code, thisworkbook_code


def generate_ta_schedule(ta_summary_df, responses_df, output_dir=None):
    """
    Generate Excel TA schedule mapping from DataFrames (optimized - no temp file I/O required)
    
    Args:
        ta_summary_df: DataFrame containing TA summary data with columns: 'TA Name', 'Hours Assigned', 'Labs Assigned'
        responses_df: DataFrame containing responses data (already loaded from responses.csv)  
        output_dir: Optional output directory (defaults to current directory)
        
    Returns:
        tuple: (excel_file_path, vba_module_file_path, vba_thisworkbook_file_path)
    """
    if output_dir is None:
        output_dir = os.getcwd()
    
    # Use the DataFrames directly instead of reading from files
    ta_summary = ta_summary_df
    responses = responses_df
    
    # Define time slots and days
    time_slots = [f"{hour}-{hour+1}" for hour in range(8, 20)]
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define colors
    occupied_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
    available_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")  # Teal/Green
    office_hours_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Yellow
    
    # Store actual sheet names for formula references
    ta_sheet_names = {}
    
    # Process each TA
    for _, ta_row in ta_summary.iterrows():
        ta_name = ta_row['TA Name']
        
        # Create sheet for this TA (truncate name if too long for Excel and make it safe)
        sheet_name = ta_name[:31] if len(ta_name) > 31 else ta_name
        # Remove problematic characters from sheet names
        sheet_name = sheet_name.replace("'", "").replace('"', "").replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(title=sheet_name)
        
        # Store the actual sheet name for later formula use
        ta_sheet_names[ta_name] = sheet_name
        
        # Set up headers
        ws.cell(row=1, column=1, value="Time Slot")
        for col, day in enumerate(days, start=2):
            ws.cell(row=1, column=col, value=day)
        
        # Parse lab assignments for this TA
        lab_assignments = parse_lab_assignments(ta_row['Labs Assigned'])
        
        # Find corresponding response row for this TA
        response_row = responses[responses['Name'].str.strip() == ta_name]
        unavailable_slots = {}
        if not response_row.empty:
            unavailable_slots = get_ta_unavailable_slots(response_row.iloc[0], responses)
        
        # Fill the schedule grid
        for row, time_slot in enumerate(time_slots, start=2):
            # Set time slot label
            ws.cell(row=row, column=1, value=time_slot)
            
            for col, day in enumerate(days, start=2):
                cell = ws.cell(row=row, column=col)
                
                # Check if this slot is occupied by a lab assignment
                is_occupied = False
                slot_hour = int(time_slot.split('-')[0])
                
                for lab_day, start_hour, end_hour in lab_assignments:
                    if (lab_day == day and start_hour <= slot_hour < end_hour):
                        is_occupied = True
                        break
                
                # Check if this slot is unavailable
                is_unavailable = False
                if time_slot in unavailable_slots:
                    if day in unavailable_slots[time_slot]:
                        is_unavailable = True
                
                # Set cell value and color
                if is_occupied:
                    cell.value = "Occupied"
                    cell.fill = occupied_fill
                elif is_unavailable:
                    # Leave blank for unavailable slots
                    cell.value = ""
                else:
                    cell.value = "Available"
                    cell.fill = available_fill
        
        # Add summary row at the bottom
        summary_row = len(time_slots) + 3  # +2 for header row and +1 for spacing
        
        # Add summary headers
        ws.cell(row=summary_row, column=1, value="TA Name:")
        ws.cell(row=summary_row, column=2, value=ta_row['TA Name'])
        
        ws.cell(row=summary_row + 1, column=1, value="Hours Assigned:")
        ws.cell(row=summary_row + 1, column=2, value=ta_row['Hours Assigned'])
        
        ws.cell(row=summary_row + 2, column=1, value="Remaining Hours:")
        # Calculate remaining hours dynamically: Hours Assigned - Office Hours
        remaining_formula = f'=B{summary_row + 1}-B{summary_row + 4}'
        ws.cell(row=summary_row + 2, column=2, value=remaining_formula)
        
        ws.cell(row=summary_row + 3, column=1, value="Labs Assigned:")
        labs_assigned_value = ta_row['Labs Assigned'] if pd.notna(ta_row['Labs Assigned']) and ta_row['Labs Assigned'] != "None" else "None"
        ws.cell(row=summary_row + 3, column=2, value=labs_assigned_value)
        
        ws.cell(row=summary_row + 4, column=1, value="Office Hours:")
        # Use simple COUNTIF formula that definitely works
        oh_range = f"B2:F{len(time_slots) + 1}"  # B2 to F13 (for 12 time slots)
        oh_formula = f'=COUNTIF({oh_range},"OH")'
        ws.cell(row=summary_row + 4, column=2, value=oh_formula)
        
        # Add conditional formatting to color "OH" entries yellow
        oh_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        oh_rule = FormulaRule(formula=[f'B2="OH"'], fill=oh_fill)
        ws.conditional_formatting.add(oh_range, oh_rule)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create Office Hours summary sheet
    oh_ws = wb.create_sheet(title="Office Hours")
    
    # Set up headers for Office Hours sheet
    oh_ws.cell(row=1, column=1, value="Time Slot")
    for col, day in enumerate(days, start=2):
        oh_ws.cell(row=1, column=col, value=day)
    
    # Fill the Office Hours schedule grid with formulas
    for row, time_slot in enumerate(time_slots, start=2):
        # Set time slot label
        oh_ws.cell(row=row, column=1, value=time_slot)
        
        for col, day in enumerate(days, start=2):
            cell = oh_ws.cell(row=row, column=col)
            # VBA will handle this, so leave empty for now
            cell.value = ""
    
    # Add summary formulas for Office Hours sheet
    oh_summary_row = len(time_slots) + 3
    
    oh_ws.cell(row=oh_summary_row, column=1, value="Total Office Hour Slots:")
    # Count non-empty cells in the Office Hours grid
    oh_ws.cell(row=oh_summary_row, column=2, value=f"=COUNTA(B2:F{len(time_slots) + 1})")
    
    oh_ws.cell(row=oh_summary_row + 1, column=1, value="Total TAs in Office Hours:")
    # Sum up COUNTIF formulas from all TA sheets
    if ta_sheet_names:
        count_formulas = []
        for ta_name, sheet_name in ta_sheet_names.items():
            if " " in sheet_name:
                count_formulas.append(f"COUNTIF('{sheet_name}'!B2:F{len(time_slots) + 1},\"OH\")")
            else:
                count_formulas.append(f"COUNTIF({sheet_name}!B2:F{len(time_slots) + 1},\"OH\")")
        
        if count_formulas:
            formula = f"={'+'.join(count_formulas)}"
            oh_ws.cell(row=oh_summary_row + 1, column=2, value=formula)
        else:
            oh_ws.cell(row=oh_summary_row + 1, column=2, value=0)
    else:
        oh_ws.cell(row=oh_summary_row + 1, column=2, value=0)
    
    # Auto-adjust column widths for Office Hours sheet
    for column in oh_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)  # Slightly wider for multiple TA names
        oh_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate VBA code files for manual installation
    module_code, thisworkbook_code = create_vba_instructions(ta_sheet_names)
    
    # Save VBA code to files in the output directory
    vba_module_path = os.path.join(output_dir, "vba_module_code.txt")
    vba_thisworkbook_path = os.path.join(output_dir, "vba_thisworkbook_code.txt")
    
    with open(vba_module_path, "w") as f:
        f.write(module_code)
    
    with open(vba_thisworkbook_path, "w") as f:
        f.write(thisworkbook_code)
    
    # Save the workbook as regular Excel file (VBA will be added manually)
    output_filename = os.path.join(output_dir, "TA_Schedule_Mapping.xlsx")
    wb.save(output_filename)
    
    return output_filename, vba_module_path, vba_thisworkbook_path
